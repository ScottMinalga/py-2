# Author: Prof C.

# Partly sourced from https://www.geeksforgeeks.org/working-with-excel-spreadsheets-in-python/

# Python program to read an excel file 
  
# import openpyxl module 
import openpyxl 
  
# Give the location of the file 
path = "SampleExcel02.xlsx"
  
# To open the workbook 
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
  
# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 
  
# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column
  
print("Total Rows:", row)
print("Total Columns:", column)
  
# printing the value of first column
# Loop will print all values 
# of first column  
print("\nValue of first column")
for i in range(1, row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 1) 
    print(cell_obj.value) 
      
# printing the value of first column
# Loop will print all values 
# of first row
print("\nValue of first row")
for i in range(1, column + 1):
    cell_obj = sheet_obj.cell(row = 2, column = i)
    print(cell_obj.value, end = " ")
